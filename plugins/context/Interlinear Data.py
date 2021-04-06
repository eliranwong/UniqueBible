import config
from qtpy.QtWidgets import QWidget


class InterlinearDataWindow(QWidget):

    translation = (
        "Interlinear Data",
        "Enter a reference:",
        "No bible verse reference is found!",
        "Export to Spreadsheet",
        "Close",
    )

    def __init__(self, parent, initialVerse=""):
        super().__init__()
        self.parent = parent
        # set title
        self.setWindowTitle(self.translation[0])
        self.setMinimumSize(830, 500)
        # set variables
        self.setupVariables()
        # setup interface
        self.setupUI(initialVerse)

    def setupVariables(self):
        import os, sqlite3
        # connect bibles.sqlite
        self.database = os.path.join(config.marvelData, "morphology.sqlite")
        self.connection = sqlite3.connect(self.database)
        self.cursor = self.connection.cursor()

    def __del__(self):
        self.connection.close()

    def setupUI(self, initialVerse=""):
        from qtpy.QtGui import QStandardItemModel
        from qtpy.QtWidgets import (QPushButton, QLabel, QTableView, QHBoxLayout, QVBoxLayout, QLineEdit)
        #from qtpy.QtWidgets import QAbstractItemView

        mainLayout = QVBoxLayout()

        mainLayout.addWidget(QLabel(self.translation[0]))

        layout = QHBoxLayout()
        layout.addWidget(QLabel(self.translation[1]))
        self.searchEntry = QLineEdit()
        # Set initial entry
        self.searchEntry.setText(initialVerse if initialVerse else "John 3:16")
        self.searchEntry.returnPressed.connect(self.resetItems)
        layout.addWidget(self.searchEntry)
        mainLayout.addLayout(layout)

        self.dataView = QTableView()
        # Allow editing so that users can select text and copy
        #self.dataView.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.dataView.setSortingEnabled(True)
        self.dataViewModel = QStandardItemModel(self.dataView)
        self.dataView.setModel(self.dataViewModel)
        self.resetItems()
        mainLayout.addWidget(self.dataView)

        buttonLayout = QHBoxLayout()
        button = QPushButton(self.translation[4])
        button.clicked.connect(self.close)
        buttonLayout.addWidget(button)
        button = QPushButton(self.translation[3])
        button.clicked.connect(self.exportSpreadsheet)
        buttonLayout.addWidget(button)
        mainLayout.addLayout(buttonLayout)

        self.setLayout(mainLayout)

    def resetItems(self):
        from BibleVerseParser import BibleVerseParser
        from qtpy.QtGui import QStandardItem

        # Empty the model before reset
        self.dataViewModel.clear()
        # Reset
        # Parse entered reference
        reference = self.searchEntry.text().strip()
        verses = BibleVerseParser(config.parserStandarisation).extractAllReferences(reference, False)
        if verses:
            # Search morphology database
            bcv = verses[0][0:3]
            query = "SELECT * FROM morphology WHERE Book = ? AND Chapter = ? AND Verse = ?"
            self.cursor.execute(query, bcv)
            self.results = self.cursor.fetchall()
            # Display data
            # TABLE morphology (WordID INT, ClauseID INT, Book INT, Chapter INT, Verse INT, Word TEXT, LexicalEntry TEXT, MorphologyCode TEXT, Morphology TEXT, Lexeme TEXT, Transliteration TEXT, Pronunciation TEXT, Interlinear TEXT, Translation TEXT, Gloss TEXT)
            #for wordID, clauseID, b, c, v, textWord, lexicalEntry, morphologyCode, morphology, lexeme, transliteration, pronuciation, interlinear, translation, gloss in self.results:
            for row, result in enumerate(self.results):
                for column in range(0, len(result)):
                    text = str(result[column]) if column < 5 else result[column]
                    item = QStandardItem(text)
                    self.dataViewModel.setItem(row, column, item)
            self.dataViewModel.setHorizontalHeaderLabels(["WordID", "ClauseID", "Book", "Chapter", "Verse", "Word", "LexicalEntry", "MorphologyCode", "Morphology", "Lexeme", "Transliteration", "Pronunciation", "Interlinear", "Translation", "Gloss"])
        else:
            self.results = []
            self.displayMessage(self.translation[2])
        self.dataView.resizeColumnsToContents()

    def displayMessage(self, message):
        from qtpy.QtWidgets import QMessageBox
        QMessageBox.information(self, self.translation[0], message)

    def exportSpreadsheet(self):
        import sys
        from install.module import installmodule

        module = "openpyxl"

        # Check if essential module is installed
        try:
            from openpyxl import Workbook
            openpyxlInstalled = True
        except:
            openpyxlInstalled = False

        # Install essential module if it is absent
        if not openpyxlInstalled:
            installmodule(module)
        if not module in sys.modules:
            try:
                from pypinyin import pinyin
                self.runExportSpreadsheet()
            except:
                self.displayMessage("Package '{0}' is required but not installed!\nRun 'pip3 install {0}' to install it first.".format(module))
        else:
            self.runExportSpreadsheet()

    def runExportSpreadsheet(self):
        import os, platform, subprocess
        from openpyxl import Workbook
        from openpyxl.styles import Font

        # Specify excel file path
        filePath = os.path.join("plugins", "context", "Interlinear Data.xlsx")

        # Documentation on openpyxl: https://openpyxl.readthedocs.io/en/stable/
        wb = Workbook()
        # grab the active worksheet
        ws = wb.active
        # Append rows
        ws.append(["WordID", "ClauseID", "Book", "Chapter", "Verse", "Word", "LexicalEntry", "MorphologyCode", "Morphology", "Lexeme", "Transliteration", "Pronunciation", "Interlinear", "Translation", "Gloss"])
        if self.results:
            for result in self.results:
                ws.append(result)
            # Apply style
            # Documentation: https://openpyxl.readthedocs.io/en/stable/styles.html
            font = Font(name="Calibri") if self.results[1][2] >= 40 else Font(name="Ezra SIL")
            for column in ("F", "J"):
                for row in range(0, len(self.results)):
                    ws["{0}{1}".format(column, row + 2)].font = font
            font = Font(name="Calibri")
            for column in ("K", "L"):
                for row in range(0, len(self.results)):
                    ws["{0}{1}".format(column, row + 2)].font = font

        # Save the file
        wb.save(filePath)
        if platform.system() == "Linux":
            subprocess.Popen([config.open, filePath])
        else:
            os.system("{0} {1}".format(config.open, filePath))

config.mainWindow.bibleReadingPlan = InterlinearDataWindow(config.mainWindow, config.pluginContext)
config.mainWindow.bibleReadingPlan.show()
