from uniquebible import config
if config.qtLibrary == "pyside6":
    from PySide6.QtWidgets import QWidget
else:
    from qtpy.QtWidgets import QWidget


class InterlinearDataWindow(QWidget):

    translation = (
        "Interlinear Data",
        "Enter a reference:",
        "No bible verse reference is found!",
        "Export to Spreadsheet",
        "Close",
        "Add to Workspace",
    )

    def __init__(self, parent, initialVerse=""):
        super().__init__()
        self.parent = parent
        # set title
        self.setWindowTitle(self.translation[0])
        self.setMinimumSize(830, 500)
        # set variables
        self.setupVariables()
        # setup interface
        self.setupUI(initialVerse)

    def setupVariables(self):
        import os, sqlite3
        # connect bibles.sqlite
        self.database = os.path.join(config.marvelData, "morphology.sqlite")
        self.connection = sqlite3.connect(self.database)
        self.cursor = self.connection.cursor()
        # Spreadsheet headers
        self.headers = ("WordID", "ClauseID", "Book", "Chapter", "Verse", "Word", "LexicalEntry", "MorphologyCode", "Morphology", "Lexeme", "Transliteration", "Pronunciation", "Interlinear", "Translation", "Gloss")

    def __del__(self):
        self.connection.close()

    def setupUI(self, initialVerse=""):
        if config.qtLibrary == "pyside6":
            from PySide6.QtGui import QStandardItemModel
            from PySide6.QtWidgets import QPushButton, QLabel, QTableView, QHBoxLayout, QVBoxLayout, QLineEdit
        else:
            from qtpy.QtGui import QStandardItemModel
            from qtpy.QtWidgets import QPushButton, QLabel, QTableView, QHBoxLayout, QVBoxLayout, QLineEdit
        #from qtpy.QtWidgets import QAbstractItemView

        mainLayout = QVBoxLayout()

        mainLayout.addWidget(QLabel(self.translation[0]))

        layout = QHBoxLayout()
        layout.addWidget(QLabel(self.translation[1]))
        self.searchEntry = QLineEdit()
        self.searchEntry.setClearButtonEnabled(True)
        # Set initial entry
        self.searchEntry.setText(initialVerse if initialVerse else self.parent.bcvToVerseReference(config.mainB, config.mainC, config.mainV))
        self.searchEntry.returnPressed.connect(self.resetItems)
        layout.addWidget(self.searchEntry)
        mainLayout.addLayout(layout)

        self.dataView = QTableView()
        # Allow editing so that users can select text and copy
        #self.dataView.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.dataView.setSortingEnabled(True)
        self.dataViewModel = QStandardItemModel(self.dataView)
        self.dataView.setModel(self.dataViewModel)
        self.resetItems()
        mainLayout.addWidget(self.dataView)

        buttonLayout = QHBoxLayout()
        button = QPushButton(self.translation[4])
        button.clicked.connect(self.close)
        buttonLayout.addWidget(button)
        button = QPushButton(self.translation[3])
        button.clicked.connect(self.exportSpreadsheet)
        buttonLayout.addWidget(button)
        if ("Tabulate" in config.enabled):
            button = QPushButton(self.translation[5])
            button.clicked.connect(self.addToWorkspace)
            buttonLayout.addWidget(button)
        mainLayout.addLayout(buttonLayout)

        self.setLayout(mainLayout)

    def resetItems(self):
        from uniquebible.util.BibleVerseParser import BibleVerseParser
        if config.qtLibrary == "pyside6":
            from PySide6.QtGui import QStandardItem
        else:
            from qtpy.QtGui import QStandardItem

        # Empty the model before reset
        self.dataViewModel.clear()
        # Reset
        # Parse entered reference
        reference = self.searchEntry.text().strip()
        verses = BibleVerseParser(config.parserStandarisation).extractAllReferences(reference, False)
        if verses:
            # Search morphology database
            bcv = verses[0][0:3]
            query = "SELECT * FROM morphology WHERE Book = ? AND Chapter = ? AND Verse = ?"
            self.cursor.execute(query, bcv)
            self.results = self.cursor.fetchall()
            # Display data
            # TABLE morphology (WordID INT, ClauseID INT, Book INT, Chapter INT, Verse INT, Word TEXT, LexicalEntry TEXT, MorphologyCode TEXT, Morphology TEXT, Lexeme TEXT, Transliteration TEXT, Pronunciation TEXT, Interlinear TEXT, Translation TEXT, Gloss TEXT)
            #for wordID, clauseID, b, c, v, textWord, lexicalEntry, morphologyCode, morphology, lexeme, transliteration, pronuciation, interlinear, translation, gloss in self.results:
            for row, result in enumerate(self.results):
                for column in range(0, len(result)):
                    text = str(result[column]) if column < 5 else result[column]
                    item = QStandardItem(text)
                    self.dataViewModel.setItem(row, column, item)
            self.dataViewModel.setHorizontalHeaderLabels(["WordID", "ClauseID", "Book", "Chapter", "Verse", "Word", "LexicalEntry", "MorphologyCode", "Morphology", "Lexeme", "Transliteration", "Pronunciation", "Interlinear", "Translation", "Gloss"])
        else:
            self.results = []
            self.displayMessage(self.translation[2])
        self.dataView.resizeColumnsToContents()

    def displayMessage(self, message):
        if config.qtLibrary == "pyside6":
            from PySide6.QtWidgets import QMessageBox
        else:
            from qtpy.QtWidgets import QMessageBox
        QMessageBox.information(self, self.translation[0], message)

    def addToWorkspace(self):
        from tabulate import tabulate
        results = self.results
        results.insert(0, ("WordID", "ClauseID", "Book", "Chapter", "Verse", "Word", "LexicalEntry", "MorphologyCode", "Morphology", "Lexeme", "Transliteration", "Pronunciation", "Interlinear", "Translation", "Gloss"))
        html = tabulate(results, tablefmt='html')
        config.mainWindow.addToWorkspaceReadOnlyAction(html, self.translation[0])

    def exportSpreadsheet(self):
        import sys
        from uniquebible.install.module import installmodule

        module = "openpyxl"

        # Check if essential module is installed
        try:
            from openpyxl import Workbook
            moduleInstalled = True
        except:
            moduleInstalled = False

        # Install essential module if it is absent
        if not moduleInstalled:
            installmodule(module)
        if not module in sys.modules:
            try:
                from openpyxl import Workbook
                self.runExportSpreadsheet()
            except:
                #self.displayMessage("Package '{0}' is required but not installed!\nRun 'pip3 install {0}' to install it first.".format(module))
                # openpyxl requires pyton version 3.6+, try alternative 'xlsxwriter'
                self.exportSpreadsheet2()
        else:
            self.runExportSpreadsheet()

    def runExportSpreadsheet(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font

        # Specify excel file path
        #filePath = os.path.join(os.getcwd(), "plugins", "menu", "Interlinear_Data.xlsx")
        filePath = self.getFilePath()
        if filePath:

            # Documentation on openpyxl: https://openpyxl.readthedocs.io/en/stable/
            wb = Workbook()
            # grab the active worksheet
            ws = wb.active
            ws.title = "UniqueBible.app"
            # Append rows
            ws.append(self.headers)
            font = Font(bold=True)
            for i in range(0, len(self.headers)):
                # row and column number starts from 1 when calling ws.cell
                ws.cell(row=1, column=i + 1).font = font
            if self.results:
                for result in self.results:
                    ws.append(result)
                # Apply style
                # Documentation: https://openpyxl.readthedocs.io/en/stable/styles.html
                font = Font(name="Calibri") if self.results[0][2] >= 40 else Font(name="Ezra SIL")
                for column in ("F", "J"):
                    for row in range(0, len(self.results)):
                        ws["{0}{1}".format(column, row + 2)].font = font
                font = Font(name="Calibri")
                for column in ("K", "L"):
                    for row in range(0, len(self.results)):
                        ws["{0}{1}".format(column, row + 2)].font = font
    
            # Save and open the file
            wb.save(filePath)
            self.openFile(filePath)

    # Use 'xlsxwriter' to export excel file if 'openpyxl' is not installed.
    def exportSpreadsheet2(self):
        import sys
        from uniquebible.install.module import installmodule

        module = "xlsxwriter"

        # Check if essential module is installed
        try:
            import xlsxwriter
            moduleInstalled = True
        except:
            moduleInstalled = False

        # Install essential module if it is absent
        if not moduleInstalled:
            installmodule(module)
        if not module in sys.modules:
            try:
                import xlsxwriter
                self.runExportSpreadsheet2()
            except:
                #self.displayMessage("Package '{0}' is required but not installed!\nRun 'pip3 install {0}' to install it first.".format(module))
                self.exportSpreadsheet3()
        else:
            self.runExportSpreadsheet2()

    def runExportSpreadsheet2(self):
        import xlsxwriter

        # Specify excel file path
        #filePath = os.path.join(os.getcwd(), "plugins", "menu", "Interlinear_Data.xlsx")
        filePath = self.getFilePath()
        if filePath:

            # Create an new Excel file and add a worksheet.
            # Documentation on xlsxwriter: https://pypi.org/project/XlsxWriter/
            workbook = xlsxwriter.Workbook(filePath)
            worksheet = workbook.add_worksheet("UniqueBible.app")
            
            # Add formats to cells.
            bold = workbook.add_format({'bold': True})
            format_right_to_left = workbook.add_format({'reading_order': 2})
            
            # Text with formatting.
            for index, header in enumerate(self.headers):
                worksheet.write(0, index, header, bold)
    
            if self.results:
                for row, result in enumerate(self.results):
                    for column, item in enumerate(result):
                        if column in (5, 9) and self.results[0][2] < 40:
                            worksheet.write(row + 1, column, item, format_right_to_left)
                        else:
                            worksheet.write(row + 1, column, item)
    
            workbook.close()
    
            # Open the saved file
            self.openFile(filePath)

    # export to csv when users cannot install either openpyxl or xlsxwriter for some reasons
    def exportSpreadsheet3(self):
        # Define a file path
        #filePath = os.path.join(os.getcwd(), "plugins", "menu", "Interlinear_Data.csv")
        filePath = self.getFilePath("csv")
        if filePath:

            # Format data
            fileContent = '"{0}"'.format('","'.join(self.headers))
            if self.results:
                for result in self.results:
                    row = [str(item) if index < 5 else item for index, item in enumerate(result)]
                    fileContent += '\n"{0}"'.format('","'.join(row))
            # Write data into file
            with open(filePath, "w") as fileObj:
                fileObj.write(fileContent)
            self.openFile(filePath)

    def getFilePath(self, fileExtension="xlsx"):
        if config.qtLibrary == "pyside6":
            from PySide6.QtWidgets import QFileDialog
        else:
            from qtpy.QtWidgets import QFileDialog

        defaultName = "Interlinear_Data.{0}".format(fileExtension)
        options = QFileDialog.Options()
        filePath, *_ = QFileDialog.getSaveFileName(self,
                config.thisTranslation["note_saveAs"],
                defaultName,
                "Spreadsheet File (*.{0})".format(fileExtension), "", options)
        if filePath:
            filePath = filePath.replace(" ", "_")
            if not filePath.endswith(".{0}".format(fileExtension)):
                filePath = "{0}.{1}".format(filePath, fileExtension)
            return filePath
        else:
            return ""

    def openFile(self, filePath):
        import platform, subprocess, os

        if platform.system() == "Linux":
            subprocess.Popen([config.open, filePath])
        elif platform.system() == "Windows":
            try:
                subprocess("excel.exe {0}".format(filePath), shell=True)
            except:
                try:
                    subprocess("start {0}".format(filePath), shell=True)
                except:
                    pass
        else:
            os.system("{0} {1}".format(config.open, filePath))


config.mainWindow.bibleReadingPlan = InterlinearDataWindow(config.mainWindow)
config.mainWindow.bibleReadingPlan.show()
